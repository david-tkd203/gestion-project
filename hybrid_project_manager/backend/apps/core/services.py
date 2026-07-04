"""
Smart Importer — Excel-to-Database engine.

Parses the VINCULOsync Excel format:
  - Leyenda sheet → Sprints + Areas
  - Sprint N sheets → Tareas + Dependencies
Builds a directed dependency graph, auto-calculates Gantt dates,
and matches participants by name.
"""
import datetime
import logging
import re
from collections import defaultdict, deque

import openpyxl
from django.db import transaction
from django.contrib.auth import get_user_model
from django.utils.dateparse import parse_date

from .models import (
    Proyecto,
    Sprint,
    Area,
    Tarea,
)
from .models import parse_acceptance_criteria, parse_duration_range  # imported from models.py

logger = logging.getLogger(__name__)


def _parse_date(val: str):
    """Parse 'YYYY-MM-DD' string to date, or return None."""
    if not val or val in ("0", "None", ""):
        return None
    return parse_date(val)

logger = logging.getLogger(__name__)
User = get_user_model()

# ─── Sheet name → sprint codigo mapping ───
SPRINT_SHEET_MAP = {
    "Sprint 0": "SP0",
    "Sprint 1": "SP1",
    "Sprint 2": "SP2",
    "Sprint 3": "SP3",
    "Sprint 4": "SP4",
    "Sprint 5": "SP5",
    "Sprint 6": "SP6",
    "Sprint 7": "SP7",
}

# ─── Area codes from the Legend ───
AREA_MAP = {
    "dirección de proyecto": ("DIR", "#3b82f6"),
    "arquitectura": ("ARQ", "#8b5cf6"),
    "gestión cultural": ("GEST", "#f59e0b"),
    "diseño ux/ui": ("UXUI", "#ec4899"),
    "liderazgo técnico osuc": ("LIDE", "#10b981"),
    "desarrollo fullstack": ("FULL", "#06b6d4"),
    "desarrollo backend y ciberseguridad": ("BACK", "#ef4444"),
}


class ImportResult:
    """Structured result of an Excel import."""

    def __init__(self):
        self.proyecto_id = None
        self.sprints_creados = 0
        self.tareas_creadas = 0
        self.dependencias_resueltas = 0
        self.dependencias_fallidas = 0
        self.usuarios_no_encontrados = []
        self.errors = []

    @property
    def success(self):
        return len(self.errors) == 0

    def __str__(self):
        return (
            f"✅ {self.tareas_creadas} tareas, {self.sprints_creados} sprints, "
            f"{self.dependencias_resueltas} deps resueltas"
            + (f", ⚠️ {self.dependencias_fallidas} deps fallidas" if self.dependencias_fallidas else "")
            + (f", {len(self.usuarios_no_encontrados)} usuarios no match" if self.usuarios_no_encontrados else "")
        )


def _find_user_by_name(name: str):
    """Search User by matching name fragments. Returns None if not found."""
    if not name or not name.strip():
        return None
    parts = name.strip().lower().split()
    for user in User.objects.all():
        full = f"{user.first_name} {user.last_name}".lower()
        if all(p in full for p in parts):
            return user
        if user.email and user.email.lower().startswith(parts[0]):
            return user
    return None


def _get_or_create_area(nombre_area: str) -> Area:
    """Match area name → code, create if needed."""
    key = nombre_area.strip().lower()
    if key in AREA_MAP:
        codigo, color = AREA_MAP[key]
    else:
        # Generate code from initials
        words = nombre_area.strip().split()
        codigo = "".join(w[0].upper() for w in words if w)[:6]
        color = "#6b7280"

    area, _ = Area.objects.get_or_create(
        codigo=codigo, defaults={"nombre": nombre_area.strip(), "color": color}
    )
    return area


def _parse_sprint_header(ws) -> dict:
    """Extract sprint metadata from the first row."""
    row1 = [str(c.value or "") for c in ws[1]]
    header = " | ".join(row1)
    # Extract: sprint name | OE | duration | description
    parts = [p.strip() for p in header.split("|")]
    info = {"nombre": parts[0] if len(parts) > 0 else ""}
    if len(parts) > 1:
        info["oe"] = parts[1]
    if len(parts) > 2:
        info["duracion"] = parts[2]
    if len(parts) > 3:
        info["descripcion"] = parts[3]
    # Parse duration to dates if present
    if "duracion" in info:
        start, end = parse_duration_range(info["duracion"])
        info["fecha_inicio"] = start
        info["fecha_fin"] = end
    return info


def _resolve_dependency(
    dep_raw: str, tareas_by_codigo: dict, proyecto: Proyecto
) -> list[Tarea]:
    """
    Parse a dependency string like "FULL-013, UXUI-014" or "ARQ-004"
    or "FULL-001 a FULL-007" into Task instances.

    Dependencies may be specified as:
      - "ARQ-004"            → single code
      - "FULL-013, UXUI-014" → comma-separated
      - "FULL-001 a FULL-007" → range (means all SP0-FULL-001...SP0-FULL-007)
    """
    if not dep_raw or dep_raw.strip().upper() == "N/A" or dep_raw.strip() == "":
        return []

    dep_raw = dep_raw.strip()
    tasks = []

    # Case 1: range "FULL-001 a FULL-007"
    range_match = re.match(r"(\w+)-(\d+)\s+a\s+\w+-(\d+)", dep_raw, re.IGNORECASE)
    if range_match:
        prefix = range_match.group(1)
        start_num = int(range_match.group(2))
        end_num = int(range_match.group(3))
        for n in range(start_num, end_num + 1):
            code = f"SP?-{prefix}-{n:03d}"
            # We'll try both exact and wildcard matching later
            tasks.append(code)

    # Case 2: comma-separated
    parts = [p.strip() for p in dep_raw.replace("|", ",").replace(";", ",").split(",")]
    for part in parts:
        part = part.strip()
        if not part or part.upper() == "N/A":
            continue
        tasks.append(part)

    # Resolve each to actual Tarea instances
    resolved = []
    for dep_code in tasks:
        found = None
        # Try exact match
        if dep_code in tareas_by_codigo:
            found = [tareas_by_codigo[dep_code]]
        else:
            # Try wildcard: the dep might be specified as "FULL-013"
            # but the full code is "SP0-FULL-013"
            m = re.match(r"(\w+)-(\d{3,})$", dep_code)
            if m:
                prefix = m.group(1)
                num = m.group(2)
                candidates = [
                    t for c, t in tareas_by_codigo.items()
                    if c.endswith(f"-{prefix}-{num}")
                ]
                found = candidates
            # Try the SP?-prefix wildcard from range syntax
            if not found and dep_code.startswith("SP?"):
                actual_code = dep_code.replace("SP?", "", 1).lstrip("-")
                candidates = [
                    t for c, t in tareas_by_codigo.items()
                    if c.endswith(actual_code)
                ]
                found = candidates

        if found:
            resolved.extend(found)
        else:
            logger.warning("Dependency not found: %s", dep_code)

    return resolved


@transaction.atomic
def import_excel(file_path: str, proyecto_nombre: str = "VINCULOsync") -> ImportResult:
    """
    Main entry point. Parses the Excel workbook, creates/updates
    Proyecto, Sprints, Areas, and Tareas with dependency graph.
    """
    result = ImportResult()
    wb = openpyxl.load_workbook(file_path, data_only=True)

    # 1. Get or create Proyecto
    proyecto, _ = Proyecto.objects.get_or_create(
        nombre=proyecto_nombre,
        defaults={"budget": 812500.00},
    )
    result.proyecto_id = str(proyecto.id)

    # 2. Parse Leyenda → Sprints + Areas
    if "Leyenda" in wb.sheetnames:
        ws = wb["Leyenda"]
        sprint_defs = {}
        area_defs = {}
        parsing_sprints = False
        parsing_areas = False
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
            vals = [str(v).strip() if v is not None else "" for v in row]
            if "SPRINTS" in vals[0].upper():
                parsing_sprints = True
                parsing_areas = False
                continue
            if "TIPOS DE CÓDIGO" in vals[0].upper():
                parsing_sprints = False
                parsing_areas = True
                continue
            if parsing_sprints and vals[0] and vals[1] and vals[2] and vals[1].startswith("SP"):
                # Sprint | Código | Duración | Descripción | Tareas | Color
                sprint_defs[vals[1]] = {
                    "nombre": vals[0],
                    "duracion": vals[2],
                    "descripcion": vals[3],
                }
            if parsing_areas and vals[0] and vals[1] and vals[0] not in ("Prefijo", "Tipo", "ESTRUCTURA DEL CÓDIGO"):
                area_defs[vals[0]] = {"nombre": vals[1]}
        logger.info("Parsed %d sprint defs, %d area defs", len(sprint_defs), len(area_defs))

    # 3. Parse sprint sheets → create Sprints + Tareas
    #    Build a codigo→task map for dependency resolution
    tareas_by_codigo: dict[str, Tarea] = {}
    sprint_instances: dict[str, Sprint] = {}

    # 3a. First pass: create Sprints from the Legend definitions
    for codigo, sdef in sprint_defs.items():
        start, end = parse_duration_range(sdef["duracion"])
        if not start:
            start = datetime.date(2026, 1, 1)
        if not end:
            end = start + datetime.timedelta(days=13)
        orden = int(codigo.replace("SP", ""))
        sprint, _ = Sprint.objects.get_or_create(
            proyecto=proyecto,
            codigo=codigo,
            defaults={
                "nombre": sdef["nombre"],
                "descripcion": sdef["descripcion"],
                "fecha_inicio": start,
                "fecha_fin": end,
                "orden": orden,
                "color": ["#3b82f6", "#8b5cf6", "#f59e0b", "#ec4899",
                          "#10b981", "#06b6d4", "#ef4444", "#6366f1"][orden % 8],
            },
        )
        # Update dates if they changed
        if sprint.fecha_inicio != start or sprint.fecha_fin != end:
            sprint.fecha_inicio = start
            sprint.fecha_fin = end
            sprint.save(update_fields=["fecha_inicio", "fecha_fin"])
        sprint_instances[codigo] = sprint
        result.sprints_creados += 1

    # 3b. Parse each task sheet
    sheet_sprint_map = {}
    for sheet_name in SPRINT_SHEET_MAP:
        sprint_code = SPRINT_SHEET_MAP[sheet_name]
        if sheet_name not in wb.sheetnames:
            logger.warning("Sheet %s not found", sheet_name)
            continue
        sheet_sprint_map[sheet_name] = sprint_code
        ws = wb[sheet_name]
        sprint = sprint_instances.get(sprint_code)
        if not sprint:
            continue

        # Parse header row 1 for sprint metadata (dates, description)
        # header_info = _parse_sprint_header(ws)  # row 1
        # Row 2 = column headers: Código | Área | Responsable | Título | Dependencias | ...
        # Data starts at row 3

        for row_idx, row in enumerate(ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True), start=3):
            vals = [str(v).strip() if v is not None else "" for v in row]
            codigo = vals[0] if len(vals) > 0 else ""

            # Skip empty rows or header rows
            if not codigo or "Código" in codigo or not re.match(r"SP\d", codigo):
                continue

            area_nombre = vals[1] if len(vals) > 1 else ""
            responsable_nombre = vals[2] if len(vals) > 2 else ""
            titulo = vals[3] if len(vals) > 3 else ""
            dependencias_raw = vals[4] if len(vals) > 4 else ""
            participantes_data = vals[5] if len(vals) > 5 else ""
            especs = vals[6] if len(vals) > 6 else ""
            ac_raw = vals[7] if len(vals) > 7 else ""

            # ─── New columns (v2.0) ───
            tipo_tarea_raw = vals[8] if len(vals) > 8 else ""
            complejidad_raw = vals[9] if len(vals) > 9 else ""
            estado_excel_raw = vals[10] if len(vals) > 10 else ""
            completitud_raw = vals[11] if len(vals) > 11 else ""
            riesgo_raw = vals[12] if len(vals) > 12 else ""
            stack_raw = vals[13] if len(vals) > 13 else ""
            rama_raw = vals[14] if len(vals) > 14 else ""
            evidencia_raw = vals[15] if len(vals) > 15 else ""
            fecha_ini_raw = vals[16] if len(vals) > 16 else ""
            fecha_fin_raw = vals[17] if len(vals) > 17 else ""

            # ─── Map tipo_tarea ───
            TIPO_MAP = {
                "backend": Tarea.TipoTarea.BACKEND,
                "frontend": Tarea.TipoTarea.FRONTEND,
                "fullstack": Tarea.TipoTarea.FULLSTACK,
                "devops/arquitectura": Tarea.TipoTarea.DEVOPS,
                "devops-arquitectura": Tarea.TipoTarea.DEVOPS,
                "gestion": Tarea.TipoTarea.GESTION,
                "diseno": Tarea.TipoTarea.DISENO,
            }
            tipo_tarea = TIPO_MAP.get(tipo_tarea_raw.lower().replace(" ", "-"), Tarea.TipoTarea.OTROS)

            # ─── Map complejidad ───
            COMPL_MAP = {
                "baja": Tarea.Complejidad.BAJA,
                "media": Tarea.Complejidad.MEDIA,
                "alta": Tarea.Complejidad.ALTA,
                "extrema": Tarea.Complejidad.EXTREMA,
            }
            complejidad = COMPL_MAP.get(complejidad_raw.lower(), Tarea.Complejidad.MEDIA)

            # ─── Map estado_excel ───
            ESTADO_MAP = {
                "pendiente": Tarea.EstadoExcel.PENDIENTE,
                "en progreso": Tarea.EstadoExcel.EN_PROGRESO,
                "completado": Tarea.EstadoExcel.COMPLETADO,
                "bloqueado": Tarea.EstadoExcel.BLOQUEADO,
            }
            estado_excel = ESTADO_MAP.get(estado_excel_raw.lower(), Tarea.EstadoExcel.PENDIENTE)

            # ─── Parse completitud ───
            try:
                completitud_pct = int(float(completitud_raw)) if completitud_raw else 0
            except (ValueError, TypeError):
                completitud_pct = 0

            # ─── Parse dates from Excel columns (override sprint defaults) ───
            task_start = _parse_date(fecha_ini_raw)
            task_end = _parse_date(fecha_fin_raw)
            if task_start:
                start_date = task_start
            if task_end:
                end_date = task_end

            # Parse acceptance criteria
            criterios = parse_acceptance_criteria(ac_raw)

            # Find area
            area = _get_or_create_area(area_nombre)

            # Find user
            responsable = _find_user_by_name(responsable_nombre)

            if responsable_nombre and not responsable:
                result.usuarios_no_encontrados.append(responsable_nombre)

            # Parse duration from sprint dates
            # Task duration = sprint duration by default, can be overridden
            start_date = sprint.fecha_inicio
            end_date = sprint.fecha_fin

            # Calculate duration in days
            duration = (end_date - start_date).days + 1 if start_date and end_date else 14

            # Create or update task
            tarea, created = Tarea.objects.update_or_create(
                codigo=codigo,
                defaults={
                    "proyecto": proyecto,
                    "sprint": sprint,
                    "area": area,
                    "titulo": titulo,
                    "responsable": responsable,
                    "responsable_nombre": responsable_nombre,
                    "especificaciones_tecnicas": especs,
                    "criterios_aceptacion": criterios,
                    "participantes_data": {"raw": participantes_data} if participantes_data else {},
                    "fecha_inicio_estimada": start_date,
                    "fecha_fin_estimada": end_date,
                    "duracion_dias": duration,
                    "orden": row_idx,
                    # New fields (v2.0)
                    "tipo_tarea": tipo_tarea,
                    "complejidad": complejidad,
                    "estado_excel": estado_excel,
                    "completitud_pct": completitud_pct,
                    "riesgo_mitigacion": riesgo_raw,
                    "stack_especifico": stack_raw,
                    "rama_github": rama_raw,
                    "evidencia_codigo": evidencia_raw,
                },
            )
            tareas_by_codigo[codigo] = tarea
            result.tareas_creadas += 1

    # 4. SECOND PASS: resolve dependencies (must be after ALL tasks exist)
    dependencias_pendientes = {}
    for sheet_name, sprint_code in sheet_sprint_map.items():
        ws = wb[sheet_name]
        for row_idx, row in enumerate(ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True), start=3):
            vals = [str(v).strip() if v is not None else "" for v in row]
            codigo = vals[0] if len(vals) > 0 else ""
            if not codigo or not re.match(r"SP\d", codigo):
                continue
            dependencias_raw = vals[4] if len(vals) > 4 else ""
            tarea = tareas_by_codigo.get(codigo)
            if not tarea:
                continue

            deps = _resolve_dependency(dependencias_raw, tareas_by_codigo, proyecto)
            if deps:
                tarea.dependencias.set(deps)
                result.dependencias_resueltas += len(deps)
            elif dependencias_raw.strip().upper() not in ("N/A", "", "NONE"):
                # Could not resolve — store for retry
                dependencias_pendientes[codigo] = dependencias_raw
                result.dependencias_fallidas += 1

            # Check if blocked
            tarea.check_blocked()
            if tarea.is_blocked:
                tarea.save(update_fields=["is_blocked", "blocked_reason"])

    # 5. Retry unresolved dependencies (cross-sheet references)
    if dependencias_pendientes:
        logger.info("Retrying %d unresolved deps...", len(dependencias_pendientes))
        for codigo, dep_raw in list(dependencias_pendientes.items()):
            tarea = tareas_by_codigo.get(codigo)
            if not tarea:
                continue
            deps = _resolve_dependency(dep_raw, tareas_by_codigo, proyecto)
            if deps:
                tarea.dependencias.set(deps)
                result.dependencias_resueltas += len(deps)
                del dependencias_pendientes[codigo]
                result.dependencias_fallidas -= 1

    # 6. Recalculate Gantt dates for all tasks
    _recalculate_gantt_dates(proyecto, tareas_by_codigo)

    wb.close()
    return result


def _recalculate_gantt_dates(proyecto: Proyecto, tareas_by_codigo: dict[str, Tarea]):
    """
    Topological sort of the dependency graph, then forward-schedule
    each task's start date = max(end_date of all dependencies).
    """
    # Build adjacency
    tasks = list(tareas_by_codigo.values())
    in_degree = {t.id: 0 for t in tasks}
    graph = defaultdict(list)

    for t in tasks:
        for dep in t.dependencias.all():
            if dep.id in in_degree:
                graph[dep.id].append(t.id)
                in_degree[t.id] = in_degree.get(t.id, 0) + 1

    # Kahn's algorithm
    queue = deque([t.id for t in tasks if in_degree.get(t.id, 0) == 0])
    topo_order = []
    while queue:
        tid = queue.popleft()
        topo_order.append(tid)
        for neighbor in graph[tid]:
            in_degree[neighbor] -= 1
            if in_degree[neighbor] == 0:
                queue.append(neighbor)

    id_to_task = {t.id: t for t in tasks}

    # Forward pass: calculate earliest start
    for tid in topo_order:
        t = id_to_task[tid]
        earliest_start = None
        for dep in t.dependencias.all():
            dep_end = dep.fecha_fin_estimada or dep.sprint.fecha_inicio
            if earliest_start is None or dep_end > earliest_start:
                earliest_start = dep_end

        if earliest_start:
            # Start the day after the last dependency ends
            new_start = earliest_start + datetime.timedelta(days=1)
            t.fecha_inicio_estimada = new_start
            if t.duracion_dias:
                t.fecha_fin_estimada = new_start + datetime.timedelta(days=t.duracion_dias - 1)
            t.save(update_fields=["fecha_inicio_estimada", "fecha_fin_estimada"])

    # Check for cycles
    if len(topo_order) < len(tasks):
        logger.warning(
            "Dependency cycle detected: %d tasks not in topo order",
            len(tasks) - len(topo_order),
        )
